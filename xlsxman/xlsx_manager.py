import os
import numpy as np
import cv2
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage


class XlsxManager:
    """
    A class to manage Excel files using openpyxl with additional image processing capabilities using cv2.
    """
    
    def __init__(self, filepath=None):
        """
        Initialize the XlsxManager.
        
        Args:
            filepath (str, optional): Path to the Excel file. If provided, will open the file.
                                     If not provided, creates a new workbook.
        """
        if filepath and os.path.exists(filepath):
            self.workbook = openpyxl.load_workbook(filepath)
            self.filepath = filepath
        else:
            self.workbook = openpyxl.Workbook()
            self.filepath = filepath
        
        # Set active worksheet to the first sheet
        self.active_sheet = self.workbook.active
    
    def save(self, filepath=None):
        """
        Save the workbook to a file.
        
        Args:
            filepath (str, optional): Path to save the Excel file. 
                                     If not provided, uses the filepath from initialization.
        
        Returns:
            bool: True if save was successful, False otherwise.
        """
        try:
            save_path = filepath if filepath else self.filepath
            if save_path:
                self.workbook.save(save_path)
                self.filepath = save_path
                return True
            else:
                raise ValueError("No filepath provided for saving.")
        except Exception as e:
            print(f"Error saving workbook: {e}")
            return False
    
    def select_sheet(self, sheet_name):
        """
        Select a worksheet by name.
        
        Args:
            sheet_name (str): Name of the worksheet to select.
            
        Returns:
            bool: True if sheet was selected, False if sheet doesn't exist.
        """
        if sheet_name in self.workbook.sheetnames:
            self.active_sheet = self.workbook[sheet_name]
            return True
        return False
    
    def create_sheet(self, sheet_name):
        """
        Create a new worksheet.
        
        Args:
            sheet_name (str): Name for the new worksheet.
            
        Returns:
            bool: True if sheet was created, False if sheet with same name exists.
        """
        if sheet_name not in self.workbook.sheetnames:
            self.workbook.create_sheet(sheet_name)
            self.active_sheet = self.workbook[sheet_name]
            return True
        return False
    
    def write_cell(self, row, col, value):
        """
        Write a value to a cell.
        
        Args:
            row (int): Row number (1-based).
            col (int or str): Column index (1-based) or letter.
            value: Value to write to the cell.
            
        Returns:
            bool: True if write was successful.
        """
        try:
            if isinstance(col, int):
                col = get_column_letter(col)
            self.active_sheet[f"{col}{row}"] = value
            return True
        except Exception as e:
            print(f"Error writing to cell: {e}")
            return False
    
    def read_cell(self, row, col):
        """
        Read a value from a cell.
        
        Args:
            row (int): Row number (1-based).
            col (int or str): Column index (1-based) or letter.
            
        Returns:
            The value of the cell.
        """
        try:
            if isinstance(col, int):
                col = get_column_letter(col)
            return self.active_sheet[f"{col}{row}"].value
        except Exception as e:
            print(f"Error reading cell: {e}")
            return None
    
    def write_range(self, start_row, start_col, data):
        """
        Write a 2D array of data to a range starting at the specified cell.
        
        Args:
            start_row (int): Starting row number (1-based).
            start_col (int or str): Starting column index (1-based) or letter.
            data (list): 2D list of data to write.
            
        Returns:
            bool: True if write was successful.
        """
        try:
            if isinstance(start_col, str):
                # Convert column letter to index
                start_col = openpyxl.utils.column_index_from_string(start_col)
            
            for i, row_data in enumerate(data):
                for j, value in enumerate(row_data):
                    col_letter = get_column_letter(start_col + j)
                    self.active_sheet[f"{col_letter}{start_row + i}"] = value
            return True
        except Exception as e:
            print(f"Error writing range: {e}")
            return False
    
    def read_range(self, start_row, start_col, end_row, end_col):
        """
        Read a range of cells.
        
        Args:
            start_row (int): Starting row number (1-based).
            start_col (int or str): Starting column index (1-based) or letter.
            end_row (int): Ending row number (1-based).
            end_col (int or str): Ending column index (1-based) or letter.
            
        Returns:
            list: 2D list containing the values in the specified range.
        """
        try:
            if isinstance(start_col, int):
                start_col = get_column_letter(start_col)
            if isinstance(end_col, int):
                end_col = get_column_letter(end_col)
                
            data = []
            for row in self.active_sheet[f"{start_col}{start_row}":f"{end_col}{end_row}"]:
                row_data = [cell.value for cell in row]
                data.append(row_data)
            return data
        except Exception as e:
            print(f"Error reading range: {e}")
            return None
    
    def style_cell(self, row, col, font=None, fill=None, alignment=None, border=None):
        """
        Apply styling to a cell.
        
        Args:
            row (int): Row number (1-based).
            col (int or str): Column index (1-based) or letter.
            font (Font, optional): Font style.
            fill (PatternFill, optional): Cell background fill.
            alignment (Alignment, optional): Cell alignment.
            border (Border, optional): Cell border.
            
        Returns:
            bool: True if styling was successful.
        """
        try:
            if isinstance(col, int):
                col = get_column_letter(col)
            
            cell = self.active_sheet[f"{col}{row}"]
            
            if font:
                cell.font = font
            if fill:
                cell.fill = fill
            if alignment:
                cell.alignment = alignment
            if border:
                cell.border = border
                
            return True
        except Exception as e:
            print(f"Error styling cell: {e}")
            return False
    
    def insert_image(self, row, col, image_path, width=None, height=None):
        """
        Insert an image at a specified cell.
        
        Args:
            row (int): Row number (1-based).
            col (int or str): Column index (1-based) or letter.
            image_path (str): Path to the image file.
            width (int, optional): Width of the image in pixels.
            height (int, optional): Height of the image in pixels.
            
        Returns:
            bool: True if insertion was successful.
        """
        try:
            if isinstance(col, int):
                col = get_column_letter(col)
            
            img = XLImage(image_path)
            
            if width and height:
                img.width = width
                img.height = height
                
            self.active_sheet.add_image(img, f"{col}{row}")
            return True
        except Exception as e:
            print(f"Error inserting image: {e}")
            return False
    
    def process_image_with_cv2(self, image_path, output_path=None, processing_func=None):
        """
        Process an image using OpenCV before inserting into Excel.
        
        Args:
            image_path (str): Path to the image file.
            output_path (str, optional): Path to save the processed image.
            processing_func (callable, optional): Function to process the image.
                This function should take a cv2 image as input and return a processed cv2 image.
            
        Returns:
            str: Path to the processed image, or None if processing failed.
        """
        try:
            # Read the image
            img = cv2.imread(image_path)
            
            if img is None:
                raise ValueError(f"Failed to read image: {image_path}")
            
            # Apply custom processing if provided
            if processing_func and callable(processing_func):
                img = processing_func(img)
            
            # Save the processed image
            if output_path:
                cv2.imwrite(output_path, img)
                return output_path
            else:
                # If no output path is provided, save to a temporary file
                base, ext = os.path.splitext(image_path)
                temp_path = f"{base}_processed{ext}"
                cv2.imwrite(temp_path, img)
                return temp_path
        except Exception as e:
            print(f"Error processing image: {e}")
            return None
            
    def auto_fit_columns(self):
        """
        Auto-fit column widths based on content.
        
        Returns:
            bool: True if auto-fit was successful.
        """
        try:
            for col in self.active_sheet.columns:
                max_length = 0
                column = col[0].column_letter  # Get the column letter
                
                for cell in col:
                    if cell.value:
                        # Calculate the length of the cell value
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                
                # Adjust column width (with some padding)
                adjusted_width = max_length + 2
                self.active_sheet.column_dimensions[column].width = adjusted_width
            
            return True
        except Exception as e:
            print(f"Error auto-fitting columns: {e}")
            return False
    
    def get_sheet_names(self):
        """
        Get list of all sheet names in the workbook.
        
        Returns:
            list: List of sheet names.
        """
        return self.workbook.sheetnames
    
    def delete_sheet(self, sheet_name):
        """
        Delete a worksheet.
        
        Args:
            sheet_name (str): Name of the worksheet to delete.
            
        Returns:
            bool: True if sheet was deleted, False if sheet doesn't exist.
        """
        if sheet_name in self.workbook.sheetnames:
            del self.workbook[sheet_name]
            # Set active sheet to first sheet if we deleted the active one
            if not hasattr(self, 'active_sheet') or sheet_name == self.active_sheet.title:
                self.active_sheet = self.workbook.active
            return True
        return False
    
    def close(self):
        """
        Close the workbook.
        """
        self.workbook.close()


# Helper function to create an example Excel file with data and formatting
def create_example_excel(filepath, data=None):
    """
    Create an example Excel file with formatted data.
    
    Args:
        filepath (str): Path to save the Excel file.
        data (list, optional): 2D list of data to write. If None, uses sample data.
        
    Returns:
        bool: True if creation was successful.
    """
    try:
        manager = XlsxManager()
        
        # Use sample data if none provided
        if data is None:
            data = [
                ["ID", "Name", "Age", "Department", "Salary"],
                [1, "John Smith", 35, "Engineering", 85000],
                [2, "Jane Doe", 28, "Marketing", 65000],
                [3, "Michael Brown", 42, "Finance", 95000],
                [4, "Emily Johnson", 31, "Human Resources", 55000],
                [5, "David Lee", 39, "Engineering", 90000]
            ]
        
        # Write data
        manager.write_range(1, 1, data)
        
        # Style header row
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"), 
            right=Side(style="thin"), 
            top=Side(style="thin"), 
            bottom=Side(style="thin")
        )
        
        for col in range(1, len(data[0]) + 1):
            manager.style_cell(1, col, font=header_font, fill=header_fill, 
                               alignment=header_alignment, border=thin_border)
        
        # Apply borders to all data cells
        for row in range(2, len(data) + 1):
            for col in range(1, len(data[0]) + 1):
                manager.style_cell(row, col, border=thin_border)
        
        # Auto-fit columns
        manager.auto_fit_columns()
        
        # Save the file
        result = manager.save(filepath)
        manager.close()
        return result
    
    except Exception as e:
        print(f"Error creating example Excel file: {e}")
        return False


if __name__ == "__main__":
    # Example usage
    example_file = "example.xlsx"
    create_example_excel(example_file)
    print(f"Created example Excel file: {example_file}")
    
    # Example of reading and modifying an Excel file
    manager = XlsxManager(example_file)
    print(f"Sheet names: {manager.get_sheet_names()}")
    
    # Read data from a range
    data = manager.read_range(1, 1, 6, 5)
    if data:
        print("Data read successfully:")
        for row in data:
            print(row)
    
    # Insert a new row
    new_row = [6, "Sarah Wilson", 29, "Sales", 70000]
    manager.write_range(7, 1, [new_row])
    
    # Save the modified file
    manager.save("modified_example.xlsx")
    manager.close() 